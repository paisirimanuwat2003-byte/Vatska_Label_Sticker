import os
import sys
import openpyxl
from openpyxl.styles import Protection

sys.stdout.reconfigure(encoding='utf-8')

def unlock_and_clean_workbook(file_path, output_file_path=None):
    if output_file_path is None:
        output_file_path = file_path

    wb = openpyxl.load_workbook(file_path, data_only=False)
    cleared_count = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # 1. Disable sheet protection completely
        ws.protection.disable()
        ws.protection.sheet = False
        
        # Unlock all individual cells in the worksheet
        for row in ws.iter_rows():
            for cell in row:
                cell.protection = Protection(locked=False)
                
                # 2. Check and remove any "Copy to sticker" text
                val = cell.value
                if val is not None and isinstance(val, str):
                    val_clean = val.strip().lower()
                    if 'copy to sticker' in val_clean or 'copy sticker' in val_clean:
                        cell.value = None
                        cleared_count += 1

    wb.save(output_file_path)
    print(f"Unlocked & Cleaned ({cleared_count} sticker cells removed): {os.path.basename(output_file_path)}")

def batch_unlock_and_clean_folder(folder_path):
    files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx') and not f.startswith('~$')]
    print(f"Unlocking and cleaning {len(files)} files in {folder_path}...")
    for fname in files:
        file_path = os.path.join(folder_path, fname)
        unlock_and_clean_workbook(file_path, file_path)

if __name__ == "__main__":
    agilent_dir = r"D:\vatska\software\examples\Agilent"
    batch_unlock_and_clean_folder(agilent_dir)
