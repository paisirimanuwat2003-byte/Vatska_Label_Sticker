import os
import sys
import openpyxl
from openpyxl.styles import Protection

sys.stdout.reconfigure(encoding='utf-8')

def normalize_and_protect_agilent_file(file_path, output_file_path=None, protect=True, password="vatskaheng"):
    """
    Normalizes Agilent Excel workbooks so all sheets share identical cell positions for header & QC fields,
    locks all worksheets with a standard password, and leaves designated input cells unlocked.
    
    CRITICAL: Keeps data_only=False so ALL user formulas across all sheets are 100% preserved!
    """
    if output_file_path is None:
        output_file_path = file_path

    # Load workbook WITHOUT data_only so ALL formulas are preserved
    wb = openpyxl.load_workbook(file_path, data_only=False)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # -------------------------------------------------------------
        # A. Job Worksheets (e.g. 'ใบงาน Agilent_JOB 1.1', 'ใบงาน Agilent_JOB 2.1')
        # -------------------------------------------------------------
        if sheet_name.startswith("ใบงาน Agilent"):
            # Link M7/N7, M8/N8, M9/N9 using formulas so formulas pointing to N still work!
            m7_val = ws['M7'].value
            n7_val = ws['N7'].value
            if (m7_val is None or str(m7_val).strip() in ['', 'AGA']) and n7_val is not None and not str(n7_val).startswith('='):
                ws['M7'].value = n7_val
                ws['N7'].value = '=M7'
            elif m7_val is not None and not str(m7_val).startswith('='):
                ws['N7'].value = '=M7'

            m8_val = ws['M8'].value
            n8_val = ws['N8'].value
            if (m8_val is None or str(m8_val).strip() == '') and n8_val is not None and not str(n8_val).startswith('='):
                ws['M8'].value = n8_val
                ws['N8'].value = '=M8'
            elif m8_val is not None and not str(m8_val).startswith('='):
                ws['N8'].value = '=M8'

            m9_val = ws['M9'].value
            n9_val = ws['N9'].value
            if (m9_val is None or str(m9_val).strip() == '') and n9_val is not None and not str(n9_val).startswith('='):
                ws['M9'].value = n9_val
                ws['N9'].value = '=M9'
            elif m9_val is not None and not str(m9_val).startswith('='):
                ws['N9'].value = '=M9'

            # Sheet Protection
            if protect:
                ws.protection.sheet = True
                if password:
                    ws.protection.password = password
                
                # Unlock designated editable fields in identical cell positions
                unlocked_cells = ['B7', 'F7', 'M7', 'F8', 'M8', 'F9', 'M9']
                for cell_ref in unlocked_cells:
                    ws[cell_ref].protection = Protection(locked=False)

        # -------------------------------------------------------------
        # B. QC Sheets ('P2-QC กล่อง' and 'A2-QC ถุง')
        # -------------------------------------------------------------
        elif 'QC' in sheet_name:
            if protect:
                ws.protection.sheet = True
                if password:
                    ws.protection.password = password
                
                # Unlock designated QC editable cells
                qc_unlocked_cells = ['G7', 'B8', 'B10', 'G8', 'G9', 'G10', 'G11', 'G12']
                for cell_ref in qc_unlocked_cells:
                    ws[cell_ref].protection = Protection(locked=False)

        # -------------------------------------------------------------
        # C. Any other sheets
        # -------------------------------------------------------------
        else:
            if protect:
                ws.protection.sheet = True
                if password:
                    ws.protection.password = password

    wb.save(output_file_path)
    print(f"Protected & Preserved Formulas: {os.path.basename(output_file_path)}")

def batch_process_agilent_folder(folder_path, protect=True, password="vatskaheng"):
    files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx') and not f.startswith('~$')]
    print(f"Processing {len(files)} files in {folder_path}...")
    for fname in files:
        file_path = os.path.join(folder_path, fname)
        normalize_and_protect_agilent_file(file_path, file_path, protect=protect, password=password)

if __name__ == "__main__":
    agilent_dir = r"D:\vatska\software\examples\Agilent"
    batch_process_agilent_folder(agilent_dir, protect=True, password="vatskaheng")
